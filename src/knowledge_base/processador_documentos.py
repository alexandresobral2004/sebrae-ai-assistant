import os
from typing import List, Dict, Optional
from pypdf import PdfReader
from docx import Document
from openpyxl import load_workbook
import yake
from sentence_transformers import SentenceTransformer
from langchain_text_splitters import RecursiveCharacterTextSplitter

class ProcessadorDocumentos:
    """Processa diferentes tipos de documentos e extrai palavras-chave."""
    
    def __init__(self):
        """Inicializa o processador com os modelos necessários."""
        self.text_splitter = RecursiveCharacterTextSplitter(
            chunk_size=1000,
            chunk_overlap=200,
            length_function=len
        )
        
        # Inicializa os modelos para extração de palavras-chave
        self.keyword_extractor = yake.KeywordExtractor(
            lan="pt",
            n=2,
            dedupLim=0.9,
            dedupFunc='seqm',
            windowsSize=1,
            top=20,
            features=None
        )
        self.sentence_model = SentenceTransformer('distiluse-base-multilingual-cased-v2')
        
    def extrair_palavras_chave(self, texto: str, num_palavras: int = 5) -> List[str]:
        """
        Extrai palavras-chave do texto usando YAKE.
        
        Args:
            texto: Texto para extrair palavras-chave
            num_palavras: Número de palavras-chave a extrair
            
        Returns:
            Lista de palavras-chave
        """
        keywords = self.keyword_extractor.extract_keywords(texto)
        # YAKE retorna tuplas (palavra, pontuação) - quanto menor a pontuação, melhor
        keywords.sort(key=lambda x: x[1])  # Ordena por relevância
        return [k[0] for k in keywords[:num_palavras]]
    
    def processar_pdf(self, caminho: str) -> Dict[str, any]:
        """Processa arquivo PDF."""
        try:
            reader = PdfReader(caminho)
            texto = ""
            for page in reader.pages:
                texto += page.extract_text() + "\n"
            return self._processar_texto(texto, caminho)
        except Exception as e:
            print(f"Erro ao processar PDF {caminho}: {str(e)}")
            return None
            
    def processar_docx(self, caminho: str) -> Dict[str, any]:
        """Processa arquivo Word."""
        try:
            doc = Document(caminho)
            texto = "\n".join([paragraph.text for paragraph in doc.paragraphs])
            return self._processar_texto(texto, caminho)
        except Exception as e:
            print(f"Erro ao processar DOCX {caminho}: {str(e)}")
            return None
            
    def processar_xlsx(self, caminho: str) -> Dict[str, any]:
        """Processa arquivo Excel."""
        try:
            wb = load_workbook(caminho, data_only=True)
            texto = ""
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    texto += " ".join([str(cell) if cell is not None else "" for cell in row]) + "\n"
            return self._processar_texto(texto, caminho)
        except Exception as e:
            print(f"Erro ao processar XLSX {caminho}: {str(e)}")
            return None
    
    def _processar_texto(self, texto: str, caminho: str) -> Dict[str, any]:
        """
        Processa o texto extraído, gerando chunks e palavras-chave.
        
        Args:
            texto: Texto extraído do documento
            caminho: Caminho do arquivo
            
        Returns:
            Dicionário com chunks de texto e metadados
        """
        # Extrai palavras-chave do texto completo
        palavras_chave = self.extrair_palavras_chave(texto)
        
        # Divide o texto em chunks
        chunks = self.text_splitter.split_text(texto)
        
        # Cria chunks de documento com metadados
        nome_arquivo = os.path.basename(caminho)
        doc_chunks = []
        
        for i, chunk in enumerate(chunks):
            # Extrai palavras-chave específicas do chunk
            chunk_keywords = self.extrair_palavras_chave(chunk, num_palavras=3)
            
            # Junta as palavras-chave em uma única string
            keywords_str = ", ".join(palavras_chave + chunk_keywords)

            doc_chunks.append({
                "texto": chunk,
                "metadados": {
                    "fonte": nome_arquivo,
                    "caminho": caminho,
                    "chunk_id": i,
                    "palavras_chave": keywords_str
                }
            })
        
        return {
            "chunks": doc_chunks,
            "palavras_chave": palavras_chave
        }
    
    def processar_diretorio(self, diretorio: str, ignorar_arquivos: set = None) -> List[Dict[str, any]]:
        """
        Processa todos os documentos suportados em um diretório.
        
        Args:
            diretorio: Caminho do diretório com os documentos
            ignorar_arquivos: Um conjunto de caminhos de arquivos a serem ignorados.
            
        Returns:
            Lista de chunks de documentos com metadados
        """
        if ignorar_arquivos is None:
            ignorar_arquivos = set()
            
        todos_chunks = []
        
        for root, _, arquivos in os.walk(diretorio):
            for arquivo in arquivos:
                caminho = os.path.join(root, arquivo)
                if caminho in ignorar_arquivos:
                    continue

                resultado = None
                
                try:
                    if arquivo.lower().endswith('.pdf'):
                        resultado = self.processar_pdf(caminho)
                    elif arquivo.lower().endswith('.docx'):
                        resultado = self.processar_docx(caminho)
                    elif arquivo.lower().endswith('.xlsx'):
                        resultado = self.processar_xlsx(caminho)
                    
                    if resultado and 'chunks' in resultado:
                        todos_chunks.extend(resultado['chunks'])
                        print(f"Processado: {arquivo} - {len(resultado['chunks'])} chunks")
                except Exception as e:
                    print(f"Erro ao processar {arquivo}: {str(e)}")
                    continue
        
        return todos_chunks